from __future__ import annotations

from pathlib import Path

from .base import BaseExtractor, ExtractionResult, ExtractorError


def _peek_magic(path: Path, n: int = 8) -> bytes:
    try:
        with open(path, "rb") as f:
            return f.read(n)
    except OSError:
        return b""


def _is_zip(magic: bytes) -> bool:
    # 정상 ZIP, 빈 ZIP, spanned ZIP 모두 cover.
    return magic.startswith(b"PK\x03\x04") or magic.startswith(b"PK\x05\x06") or magic.startswith(b"PK\x07\x08")


def _is_ole(magic: bytes) -> bool:
    return magic.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1")


class XlsxExtractor(BaseExtractor):
    name = "xlsx"
    extensions = (".xlsx", ".xlsm")

    def extract(self, path: Path) -> ExtractionResult:
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ExtractorError(f"openpyxl not installed: {e}")

        magic = _peek_magic(path)
        if not _is_zip(magic):
            # 확장자만 xlsx인 가짜 파일 (HTML 표, 손상 ZIP 등). 무리한 시도 안 한다.
            raise ExtractorError("not a valid XLSX (zip header missing)")

        try:
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        except Exception as e:
            raise ExtractorError(f"XLSX open failed: {e}")

        parts: list[str] = []
        try:
            for sheet in wb.worksheets:
                parts.append(f"# {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None and str(c).strip()]
                    if cells:
                        parts.append("\t".join(cells))
        finally:
            wb.close()

        return ExtractionResult(text="\n".join(parts))


class XlsExtractor(BaseExtractor):
    name = "xls"
    extensions = (".xls",)

    def extract(self, path: Path) -> ExtractionResult:
        try:
            import xlrd
        except ImportError as e:
            raise ExtractorError(f"xlrd not installed: {e}")

        # 확장자가 .xls지만 실제로는 .xlsx(ZIP) 인 파일이 흔하다.
        # 그런 파일은 xlrd에 넘기면 깨지므로 openpyxl로 라우팅한다.
        magic = _peek_magic(path)
        if _is_zip(magic):
            return XlsxExtractor().extract(path)
        # 텍스트/XML/HTML로 저장된 가짜 .xls 도 흔하다. 매직이 OLE가 아니면 컷.
        if not _is_ole(magic):
            raise ExtractorError("not a real XLS (BIFF/OLE header missing)")

        try:
            book = xlrd.open_workbook(str(path))
        except Exception as e:
            raise ExtractorError(f"XLS open failed: {e}")

        parts: list[str] = []
        for sheet in book.sheets():
            parts.append(f"# {sheet.name}")
            for row_idx in range(sheet.nrows):
                row = sheet.row_values(row_idx)
                cells = [str(c) for c in row if c not in (None, "") and str(c).strip()]
                if cells:
                    parts.append("\t".join(cells))

        return ExtractionResult(text="\n".join(parts))
